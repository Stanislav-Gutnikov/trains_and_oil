from datetime import datetime
from sqlalchemy import select
from sqlalchemy.orm import Session
import openpyxl
import os
from dotenv import load_dotenv

from app.core.terminal import Terminal
from app.core.train import Train
from app.db.db import Base
from app.db.models.raduzhny import Raduzhny
from app.db.models.zvezda import Zvezda
from app.db.models.polarny import Polarny


load_dotenv()


class CRUD:

    def __init__(self):
        self.model_list = [Raduzhny, Zvezda, Polarny]

    def get_obj_info_pp(self, db_object, column_index: int):
        info_dict = {
            1: str(db_object.datetime),
            2: db_object.oil,
            3: db_object.train_name,
            4: db_object.train_unloading,
            5: db_object.calc_id
        }
        return info_dict.get(column_index)

    def get_obj_info_tp(self, db_object, column_index: int):
        info_dict = {
            1: str(db_object.datetime),
            2: db_object.oil,
            3: db_object.train_name_1,
            4: db_object.train_1_unloading,
            5: db_object.train_name_2,
            6: db_object.train_2_unloading,
            7: db_object.train_name_3,
            8: db_object.train_3_unloading,
            9: db_object.calc_id
        }
        return info_dict.get(column_index)

    def get(
            self,
            calc_id: int,
            session: Session
    ):
        excel_file_path = os.getenv('XLSX_FILE_PATH')
        f = 0
        db_table_names = list(Base.metadata.tables.keys())
        wb = openpyxl.Workbook()
        for table in db_table_names:
            print(table)
            wb.create_sheet(table)
        wb.remove(wb['Sheet'])
        wb.save(excel_file_path)
        for model in self.model_list:
            obj = session.execute(select(model).where(
                model.calc_id == calc_id
            ))
            obj = obj.scalars().all()
            sheet = wb[db_table_names[f]]
            print(wb.index(sheet))
            for j in range(1, 337):
                for k in range(1, 10):
                    cell = sheet.cell(row=j, column=k)
                    if f <= 1:
                        cell.value = self.get_obj_info_pp(obj[0], k)
                    else:
                        cell.value = self.get_obj_info_tp(obj[0], k)
                obj.pop(0)
            f += 1
            wb.save(excel_file_path)
        return excel_file_path

    def get_train_name(
            self,
            train: Train
    ):
        train_name = None
        if train is not None:
            train_name = train.name
        return train_name

    def crud_transshipment_point(
            self,
            terminal: Terminal,
            start_date: datetime,
            calc_id: int,
            session: Session
    ):
        new_obj = Polarny(
            datetime=start_date,
            oil=terminal.oil,
            train_name_1=self.get_train_name(terminal.ways.get(1)),
            train_1_unloading=terminal.unloading,
            train_name_2=self.get_train_name(terminal.ways.get(2)),
            train_2_unloading=terminal.unloading,
            train_name_3=self.get_train_name(terminal.ways.get(3)),
            train_3_unloading=terminal.unloading,
            calc_id=calc_id
        )
        session.add(new_obj)
        session.commit()
        return

    def crud_production_point(
            self,
            terminal: Terminal,
            start_date: datetime,
            calc_id: int,
            session: Session
    ):
        if terminal.name == 'raduzhny':
            new_obj = Raduzhny(
                datetime=start_date,
                oil=terminal.oil,
                train_name=self.get_train_name(terminal.ways.get(1)),
                train_unloading=terminal.unloading,
                calc_id=calc_id
            )
        else:
            new_obj = Zvezda(
                datetime=start_date,
                oil=terminal.oil,
                train_name=self.get_train_name(terminal.ways.get(1)),
                train_unloading=terminal.unloading,
                calc_id=calc_id
            )
        session.add(new_obj)
        session.commit()
        return

    def post_to_db(
        self,
        terminal: Terminal,
        start_date: datetime,
        calc_id: int,
        session: Session
    ):
        if terminal.type != 'transshipment_point':
            self.crud_production_point(
                terminal,
                start_date,
                calc_id,
                session
                )
        else:
            self.crud_transshipment_point(
                terminal,
                start_date,
                calc_id,
                session
                )
        return

    def update_data(
        self,
        terminal: Terminal,
        start_date: datetime,
        calc_id: int,
        session: Session
    ):
        for model in self.model_list:
            session.query(model).filter(
                model.calc_id == calc_id,
                model.datetime == start_date
            ).delete()
        session.commit()
        self.post_to_db(terminal, start_date, calc_id, session)
        return

    def delete_obj(
            self,
            calc_id: int,
            session: Session
    ):
        for model in self.model_list:
            session.query(model).filter(
                model.calc_id == calc_id
            ).delete()
        session.commit()
        return

    def delete_all(
            self,
            session: Session
            ):
        for model in self.model_list:
            session.query(model).delete()
        session.commit()
        return


crud = CRUD()
